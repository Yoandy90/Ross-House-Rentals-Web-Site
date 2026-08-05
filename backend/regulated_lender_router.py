"""
Regulated Lender Router — Texas Chapter 342 (Subchapter E + F)
Manages:
  - Subchapter F: Small loans ≤$1,800 (15-20% monthly)
  - Subchapter E: Installment loans $500-$10,000 (tiered APR 18-30%)
  - Tax Refund Advances: Loans repaid from IRS refund
  - Reports & Excel Export
  - Collection Notifications (SMS/Email reminders)
"""
import os
import io
import base64
import logging
from datetime import datetime, timedelta
from typing import Optional, List
from fastapi import APIRouter, HTTPException, Request, Query
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import shared business logic (calculations, schedule, PDF mapping)
from loan_shared_service import (
    calculate_subchapter_f,
    calculate_subchapter_e,
    calculate_hybrid,
    calculate_tax_advance,
    generate_regulated_schedule,
    build_pdf_loan_data,
    generate_schedule_for_loan,
)

logger = logging.getLogger(__name__)

regulated_lender_router = APIRouter()

_db = None
_get_user_from_token = None
_notification_service = None


def init_regulated_lender(db_instance, get_user_func, notification_service=None):
    global _db, _get_user_from_token, _notification_service
    _db = db_instance
    _get_user_from_token = get_user_func
    _notification_service = notification_service
    logger.info("✅ Regulated Lender Router initialized")


COLLECTION_NAME = "regulated_loans"
PAYMENTS_COLLECTION = "regulated_loan_payments"


async def _auth_admin(request: Request):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = await _get_user_from_token(token)
    if not user or user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="No autorizado")
    return user


# ═══════════════════════════════════════════════════════════════════════════════
# INTEREST CALCULATION ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
# NOTE: Calculation functions (calculate_subchapter_f, calculate_subchapter_e,
# calculate_tax_advance, generate_regulated_schedule) are now imported from
# loan_shared_service.py to eliminate duplication with client_loans_router.py.
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# PYDANTIC MODELS
# ═══════════════════════════════════════════════════════════════════════════════

class LoanCreate(BaseModel):
    client_name: str
    client_phone: str = ""
    client_email: str = ""
    client_ssn_last4: str = ""
    loan_type: str  # "subchapter_f", "subchapter_e", "tax_advance"
    amount: float
    term_months: int = 1
    purpose: str = ""
    notes: str = ""
    # Tax advance specific
    expected_refund: float = 0
    tax_year: str = ""
    refund_transfer_provider: str = ""
    first_payment_date: str = ""


class PaymentCreate(BaseModel):
    amount: float
    payment_method: str = "cash"
    notes: str = ""
    source: str = "manual"  # "manual", "refund_transfer", "ach"


class StatusUpdate(BaseModel):
    status: str
    comment: str = ""


class CalculateRequest(BaseModel):
    loan_type: str
    amount: float
    term_months: int = 1


class CollectionNotificationRequest(BaseModel):
    loan_ids: List[str] = []  # Empty = all overdue
    method: str = "sms"  # "sms", "email", "both"
    message_template: str = ""  # Custom message (optional)


# ═══════════════════════════════════════════════════════════════════════════════
# ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@regulated_lender_router.get("/admin/regulated-loans/stats")
async def get_stats(request: Request):
    await _auth_admin(request)
    db = _db

    pipeline = [
        {"$group": {
            "_id": "$status",
            "count": {"$sum": 1},
            "total_amount": {"$sum": "$amount"},
            "total_balance": {"$sum": "$balance"},
            "total_interest_earned": {"$sum": "$interest_paid"},
        }}
    ]
    status_stats = {}
    async for doc in db[COLLECTION_NAME].aggregate(pipeline):
        status_stats[doc["_id"]] = {
            "count": doc["count"],
            "total_amount": doc["total_amount"],
            "total_balance": doc["total_balance"],
            "total_interest_earned": doc["total_interest_earned"],
        }

    total_loans = sum(s["count"] for s in status_stats.values())
    active_loans = status_stats.get("active", {}).get("count", 0) + status_stats.get("disbursed", {}).get("count", 0)
    total_portfolio = sum(s["total_amount"] for s in status_stats.values())
    total_balance = sum(s["total_balance"] for s in status_stats.values())
    total_interest_earned = sum(s["total_interest_earned"] for s in status_stats.values())
    delinquent = status_stats.get("delinquent", {}).get("count", 0)

    # Tax advances pending refund
    tax_advances_pending = await db[COLLECTION_NAME].count_documents({
        "loan_type": "tax_advance",
        "refund_status": {"$in": ["pending", "filed"]}
    })

    # Recent loans
    recent = []
    async for loan in db[COLLECTION_NAME].find().sort("created_at", -1).limit(10):
        recent.append({
            "id": str(loan["_id"]),
            "loan_number": loan.get("loan_number", ""),
            "client_name": loan.get("client_name", ""),
            "amount": loan.get("amount", 0),
            "loan_type": loan.get("loan_type", ""),
            "status": loan.get("status", ""),
            "created_at": loan.get("created_at", ""),
        })

    # Monthly income (last 6 months)
    six_months_ago = (datetime.now() - timedelta(days=180)).isoformat()
    monthly_income = []
    income_pipeline = [
        {"$match": {"payment_date": {"$gte": six_months_ago}}},
        {"$group": {
            "_id": {"$substr": ["$payment_date", 0, 7]},
            "total": {"$sum": "$amount"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
    ]
    async for doc in db[PAYMENTS_COLLECTION].aggregate(income_pipeline):
        monthly_income.append({"month": doc["_id"], "total": doc["total"], "count": doc["count"]})

    return {
        "total_loans": total_loans,
        "active_loans": active_loans,
        "delinquent_count": delinquent,
        "delinquency_rate": round(delinquent / active_loans * 100, 1) if active_loans > 0 else 0,
        "total_portfolio": round(total_portfolio, 2),
        "total_balance": round(total_balance, 2),
        "total_interest_earned": round(total_interest_earned, 2),
        "tax_advances_pending": tax_advances_pending,
        "by_status": status_stats,
        "recent_loans": recent,
        "monthly_income": monthly_income,
    }


@regulated_lender_router.get("/admin/regulated-loans")
async def list_loans(request: Request, search: str = "", status: str = "", loan_type: str = "", limit: int = 100, skip: int = 0):
    await _auth_admin(request)
    db = _db

    query = {}
    if search:
        query["$or"] = [
            {"client_name": {"$regex": search, "$options": "i"}},
            {"client_phone": {"$regex": search, "$options": "i"}},
            {"client_email": {"$regex": search, "$options": "i"}},
            {"loan_number": {"$regex": search, "$options": "i"}},
            {"client_ssn_last4": {"$regex": search, "$options": "i"}},
        ]
    if status:
        query["status"] = status
    if loan_type:
        query["loan_type"] = loan_type

    loans = []
    async for loan in db[COLLECTION_NAME].find(query).sort("created_at", -1).skip(skip).limit(limit):
        loan["_id"] = str(loan["_id"])
        loans.append(loan)

    total = await db[COLLECTION_NAME].count_documents(query)
    return {"loans": loans, "total": total}


@regulated_lender_router.post("/admin/regulated-loans")
async def create_loan(request: Request, body: LoanCreate):
    user = await _auth_admin(request)
    db = _db

    # ═══ BLOCK: Check if client already has an active loan (by email) ═══
    if body.client_email:
        active_loan = await db[COLLECTION_NAME].find_one({
            "client_email": body.client_email.lower(),
            "status": {"$in": ["active", "delinquent", "pending_signature"]}
        })
        if active_loan:
            ln = active_loan.get("loan_number", "")
            raise HTTPException(
                status_code=409,
                detail=f"El cliente ya tiene un préstamo activo ({ln}). Debe completar el pago antes de crear otro."
            )

    # Validate minimum amount
    if body.amount < 100:
        raise HTTPException(status_code=400, detail="Monto mínimo es $100")

    # ═══ OCCC COMPLIANCE: Enforce max term limits based on subchapter ═══
    # Subchapter F (≤$1800): max 6 months (highest credit tier)
    # Subchapter E (>$1800): max 12 months
    if body.amount <= 1800 and body.term_months > 6:
        logger.warning(f"⚠️ OCCC: Term {body.term_months}m exceeds max 6m for Sub F loan ${body.amount}. Capping to 6.")
        body.term_months = 6
    elif body.amount > 1800 and body.term_months > 12:
        logger.warning(f"⚠️ OCCC: Term {body.term_months}m exceeds max 12m for Sub E loan ${body.amount}. Capping to 12.")
        body.term_months = 12

    # Calculate interest based on type — HYBRID DEFAULT (Sub F ≤$1800, Sub E >$1800)
    if body.loan_type in ("subchapter_e", "subchapter_f", "hybrid"):
        # Hybrid logic: auto-select subchapter based on amount
        calc = calculate_hybrid(body.amount, body.term_months)
        actual_subchapter = calc.get("subchapter", "F")
        body.loan_type = f"subchapter_{actual_subchapter.lower()}"
    elif body.loan_type == "tax_advance":
        calc = calculate_tax_advance(body.amount)
        calc["total_to_pay"] = calc["total_to_collect"]
        calc["monthly_payment"] = calc["total_to_collect"]  # Single payment
        calc["total_interest"] = calc["interest_1_month"]
        calc["admin_fee"] = calc["admin_fee"]
    else:
        raise HTTPException(status_code=400, detail="loan_type inválido")

    # Generate loan number
    count = await db[COLLECTION_NAME].count_documents({})
    loan_number = f"RL-{datetime.now().year}-{count + 1:04d}"

    loan_doc = {
        "loan_number": loan_number,
        "loan_type": body.loan_type,
        "subchapter": calc.get("subchapter", "F"),
        "client_name": body.client_name,
        "client_phone": body.client_phone,
        "client_email": body.client_email,
        "client_ssn_last4": body.client_ssn_last4,
        "amount": body.amount,
        "term_months": body.term_months if body.loan_type != "tax_advance" else 1,
        "interest_rate": calc.get("effective_apr", calc.get("apr_effective", calc.get("monthly_rate_pct", 0))),
        "annual_apr": calc.get("effective_apr", calc.get("apr_effective", 0)),
        "amortization_method": "simple",
        "total_interest": calc.get("total_interest", 0),
        "admin_fee": calc.get("admin_fee", 0),
        "monthly_payment": calc.get("monthly_payment", 0),
        "total_to_pay": calc.get("total_to_pay", 0),
        "purpose": body.purpose,
        "notes": body.notes,
        "status": "active",
        "balance": body.amount + calc.get("total_interest", 0) + calc.get("admin_fee", 0),
        "principal_paid": 0,
        "interest_paid": 0,
        "fees_paid": 0,
        "days_overdue": 0,
        "disbursement_date": datetime.now().isoformat(),
        "first_payment_date": body.first_payment_date or (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"),
        "next_payment_date": body.first_payment_date or (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"),
        # Tax advance fields
        "expected_refund": body.expected_refund,
        "tax_year": body.tax_year,
        "refund_transfer_provider": body.refund_transfer_provider,
        "refund_status": "pending" if body.loan_type == "tax_advance" else None,
        "refund_received_date": None,
        "refund_amount_received": 0,
        # Calculation details
        "calculation": calc,
        # Meta
        "created_at": datetime.now().isoformat(),
        "created_by": user.get("email", "admin"),
        "updated_at": datetime.now().isoformat(),
        "status_history": [{"status": "active", "date": datetime.now().isoformat(), "by": user.get("email", "admin"), "comment": "Préstamo creado y desembolsado"}],
    }

    result = await db[COLLECTION_NAME].insert_one(loan_doc)
    loan_doc["_id"] = str(result.inserted_id)
    return {"success": True, "loan": loan_doc}


@regulated_lender_router.post("/admin/regulated-loans/calculate")
async def calculate_loan(request: Request, body: CalculateRequest):
    await _auth_admin(request)

    # ═══ OCCC COMPLIANCE: Cap term months ═══
    if body.amount <= 1800 and body.term_months > 6:
        body.term_months = 6
    elif body.amount > 1800 and body.term_months > 12:
        body.term_months = 12

    if body.loan_type == "subchapter_f":
        result = calculate_subchapter_f(body.amount, body.term_months)
    elif body.loan_type == "subchapter_e":
        result = calculate_subchapter_e(body.amount, body.term_months)
    elif body.loan_type == "hybrid":
        result = calculate_hybrid(body.amount, body.term_months)
    elif body.loan_type == "tax_advance":
        result = calculate_tax_advance(body.amount)
    else:
        raise HTTPException(status_code=400, detail="loan_type inválido")

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTS ENDPOINT
# ═══════════════════════════════════════════════════════════════════════════════

@regulated_lender_router.get("/admin/regulated-loans/reports")
async def get_reports(request: Request, period: str = "all", license_type: str = "regulated"):
    """
    Get detailed lending reports.
    period: 'month', 'quarter', 'year', 'all'
    license_type: 'regulated' or 'cab'
    """
    await _auth_admin(request)
    db = _db

    # Select collection based on license type
    loan_collection = "cab_loans" if license_type == "cab" else COLLECTION_NAME
    payments_collection = "cab_payments" if license_type == "cab" else PAYMENTS_COLLECTION

    # Date filter
    date_filter = {}
    now = datetime.now()
    if period == "month":
        start = now.replace(day=1).isoformat()
        date_filter = {"created_at": {"$gte": start}}
    elif period == "quarter":
        quarter_start_month = ((now.month - 1) // 3) * 3 + 1
        start = now.replace(month=quarter_start_month, day=1).isoformat()
        date_filter = {"created_at": {"$gte": start}}
    elif period == "year":
        start = now.replace(month=1, day=1).isoformat()
        date_filter = {"created_at": {"$gte": start}}

    # 1. Portfolio Summary
    all_loans = []
    async for loan in db[loan_collection].find(date_filter):
        all_loans.append(loan)

    total_originated = sum(l.get("amount", 0) for l in all_loans)
    total_interest_charged = sum(l.get("total_interest", 0) for l in all_loans)
    total_fees_charged = sum(l.get("admin_fee", 0) for l in all_loans)
    total_collected_principal = sum(l.get("principal_paid", 0) for l in all_loans)
    total_collected_interest = sum(l.get("interest_paid", 0) for l in all_loans)
    total_collected_fees = sum(l.get("fees_paid", 0) for l in all_loans)
    total_outstanding = sum(l.get("balance", 0) for l in all_loans)

    # 2. By Loan Type
    by_type = {}
    for loan in all_loans:
        lt = loan.get("loan_type", "unknown")
        if lt not in by_type:
            by_type[lt] = {"count": 0, "originated": 0, "outstanding": 0, "interest_earned": 0, "paid_off": 0}
        by_type[lt]["count"] += 1
        by_type[lt]["originated"] += loan.get("amount", 0)
        by_type[lt]["outstanding"] += loan.get("balance", 0)
        by_type[lt]["interest_earned"] += loan.get("interest_paid", 0)
        if loan.get("status") == "paid_off":
            by_type[lt]["paid_off"] += 1

    # 3. Delinquency Analysis
    active_loans = [l for l in all_loans if l.get("status") in ("active", "disbursed")]
    overdue_loans = []
    for loan in active_loans:
        next_pay = loan.get("next_payment_date", "")
        if next_pay and next_pay < now.strftime("%Y-%m-%d"):
            days_late = (now - datetime.strptime(next_pay, "%Y-%m-%d")).days
            overdue_loans.append({
                "loan_number": loan.get("loan_number", ""),
                "client_name": loan.get("client_name", ""),
                "client_phone": loan.get("client_phone", ""),
                "balance": loan.get("balance", 0),
                "days_overdue": days_late,
                "next_payment_date": next_pay,
                "monthly_payment": loan.get("monthly_payment", 0),
            })
    overdue_loans.sort(key=lambda x: x["days_overdue"], reverse=True)

    # 4. Collection Summary
    total_overdue_amount = sum(l["balance"] for l in overdue_loans)

    # 5. Tax Advances Status
    tax_advances = [l for l in all_loans if l.get("loan_type") == "tax_advance"]
    tax_pending = [t for t in tax_advances if t.get("refund_status") in ("pending", "filed")]
    tax_received = [t for t in tax_advances if t.get("refund_status") == "received"]

    # 6. Monthly Origination Trend
    monthly_trend = {}
    for loan in all_loans:
        month_key = loan.get("created_at", "")[:7]
        if month_key:
            if month_key not in monthly_trend:
                monthly_trend[month_key] = {"count": 0, "amount": 0}
            monthly_trend[month_key]["count"] += 1
            monthly_trend[month_key]["amount"] += loan.get("amount", 0)

    return {
        "period": period,
        "summary": {
            "total_loans": len(all_loans),
            "total_originated": round(total_originated, 2),
            "total_interest_charged": round(total_interest_charged, 2),
            "total_fees_charged": round(total_fees_charged, 2),
            "total_collected_principal": round(total_collected_principal, 2),
            "total_collected_interest": round(total_collected_interest, 2),
            "total_collected_fees": round(total_collected_fees, 2),
            "total_revenue": round(total_collected_interest + total_collected_fees, 2),
            "total_outstanding": round(total_outstanding, 2),
            "avg_loan_size": round(total_originated / len(all_loans), 2) if all_loans else 0,
            "paid_off_count": len([l for l in all_loans if l.get("status") == "paid_off"]),
            "active_count": len(active_loans),
            "delinquent_count": len(overdue_loans),
            "default_count": len([l for l in all_loans if l.get("status") == "default"]),
        },
        "by_type": by_type,
        "delinquency": {
            "overdue_loans": overdue_loans[:20],
            "total_overdue_amount": round(total_overdue_amount, 2),
            "overdue_count": len(overdue_loans),
            "delinquency_rate": round(len(overdue_loans) / len(active_loans) * 100, 1) if active_loans else 0,
        },
        "tax_advances": {
            "total": len(tax_advances),
            "pending_refund": len(tax_pending),
            "refund_received": len(tax_received),
            "pending_amount": round(sum(t.get("balance", 0) for t in tax_pending), 2),
        },
        "monthly_trend": [{"month": k, **v} for k, v in sorted(monthly_trend.items())],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT ENDPOINT
# ═══════════════════════════════════════════════════════════════════════════════

@regulated_lender_router.get("/admin/regulated-loans/export-excel")
async def export_loans_excel(request: Request, status: str = "", loan_type: str = ""):
    """Export all loans to an Excel file."""
    await _auth_admin(request)
    db = _db

    query = {}
    if status:
        query["status"] = status
    if loan_type:
        query["loan_type"] = loan_type

    loans = []
    async for loan in db[COLLECTION_NAME].find(query).sort("created_at", -1):
        loans.append(loan)

    # Create workbook
    wb = Workbook()

    # ── Sheet 1: Loans ──
    ws = wb.active
    ws.title = "Préstamos"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="8B1513", end_color="8B1513", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "# Préstamo", "Cliente", "Teléfono", "Email", "SSN (últimos 4)",
        "Tipo", "Subcapítulo", "Monto", "Plazo (meses)", "Tasa (%)",
        "Interés Total", "Admin Fee", "Total a Pagar", "Pago Mensual",
        "Balance", "Principal Pagado", "Interés Pagado", "Fees Pagado",
        "Estatus", "Fecha Desembolso", "Próximo Pago", "Propósito", "Notas"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Data rows
    type_labels = {"subchapter_f": "Subch. F", "subchapter_e": "Subch. E", "tax_advance": "Adelanto Taxes"}
    status_labels = {"active": "Activo", "paid_off": "Pagado", "delinquent": "En Mora", "default": "Incobrable", "cancelled": "Cancelado"}

    for row_idx, loan in enumerate(loans, 2):
        row_data = [
            loan.get("loan_number", ""),
            loan.get("client_name", ""),
            loan.get("client_phone", ""),
            loan.get("client_email", ""),
            loan.get("client_ssn_last4", ""),
            type_labels.get(loan.get("loan_type", ""), loan.get("loan_type", "")),
            loan.get("subchapter", ""),
            loan.get("amount", 0),
            loan.get("term_months", 0),
            loan.get("interest_rate", 0),
            loan.get("total_interest", 0),
            loan.get("admin_fee", 0),
            loan.get("total_to_pay", 0),
            loan.get("monthly_payment", 0),
            loan.get("balance", 0),
            loan.get("principal_paid", 0),
            loan.get("interest_paid", 0),
            loan.get("fees_paid", 0),
            status_labels.get(loan.get("status", ""), loan.get("status", "")),
            loan.get("disbursement_date", "")[:10] if loan.get("disbursement_date") else "",
            loan.get("next_payment_date", ""),
            loan.get("purpose", ""),
            loan.get("notes", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col in (8, 11, 12, 13, 14, 15, 16, 17, 18):  # Money columns
                cell.number_format = '$#,##0.00'

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    # ── Sheet 2: Payments ──
    ws2 = wb.create_sheet("Pagos")
    pay_headers = ["# Préstamo", "Cliente", "# Pago", "Monto", "Método", "Fuente", "Fecha", "Notas", "Registrado por"]
    for col, header in enumerate(pay_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    pay_row = 2
    for loan in loans:
        loan_id = str(loan["_id"])
        async for payment in db[PAYMENTS_COLLECTION].find({"loan_id": loan_id}).sort("payment_date", 1):
            row_data = [
                loan.get("loan_number", ""),
                loan.get("client_name", ""),
                payment.get("payment_number", ""),
                payment.get("amount", 0),
                payment.get("payment_method", ""),
                payment.get("source", ""),
                payment.get("payment_date", "")[:10] if payment.get("payment_date") else "",
                payment.get("notes", ""),
                payment.get("recorded_by", ""),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws2.cell(row=pay_row, column=col, value=val)
                cell.border = thin_border
                if col == 4:
                    cell.number_format = '$#,##0.00'
            pay_row += 1

    for col in ws2.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_length + 2, 25)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Ross_Lending_Prestamos_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@regulated_lender_router.post("/admin/regulated-loans/collection-notifications")
async def send_collection_notifications(request: Request, body: CollectionNotificationRequest):
    """Send payment reminders to clients with overdue loans."""
    user = await _auth_admin(request)
    db = _db

    now = datetime.now()

    # Get overdue loans
    if body.loan_ids:
        query = {"_id": {"$in": [ObjectId(lid) for lid in body.loan_ids]}}
    else:
        # All active loans past next_payment_date
        query = {
            "status": {"$in": ["active", "disbursed", "delinquent"]},
            "next_payment_date": {"$lt": now.strftime("%Y-%m-%d")},
        }

    overdue_loans = []
    async for loan in db[COLLECTION_NAME].find(query):
        overdue_loans.append(loan)

    if not overdue_loans:
        return {
            "success": True,
            "total_notified": 0,
            "sms_sent": 0,
            "sms_failed": 0,
            "email_sent": 0,
            "email_failed": 0,
            "details": []
        }

    results = {"sms_sent": 0, "sms_failed": 0, "email_sent": 0, "email_failed": 0, "details": []}

    for loan in overdue_loans:
        client_name = loan.get("client_name", "Cliente")
        client_phone = loan.get("client_phone", "")
        client_email = loan.get("client_email", "")
        balance = loan.get("balance", 0)
        monthly_payment = loan.get("monthly_payment", 0)
        loan_number = loan.get("loan_number", "")
        next_pay = loan.get("next_payment_date", "")

        days_late = 0
        if next_pay:
            try:
                days_late = (now - datetime.strptime(next_pay, "%Y-%m-%d")).days
            except (ValueError, TypeError):
                pass

        # Default message
        if body.message_template:
            sms_msg = body.message_template.replace("{nombre}", client_name).replace("{balance}", f"${balance:,.2f}").replace("{pago}", f"${monthly_payment:,.2f}").replace("{dias}", str(days_late)).replace("{numero}", loan_number)
        else:
            sms_msg = (
                f"Ross Lending: Hola {client_name.split()[0]}, "
                f"tu pago de ${monthly_payment:,.2f} está vencido hace {days_late} día(s). "
                f"Balance actual: ${balance:,.2f}. "
                f"Préstamo #{loan_number}. "
                f"Comunícate con nosotros al (806) 503-7721 para hacer tu pago. Gracias."
            )

        notification_record = {
            "loan_id": str(loan["_id"]),
            "loan_number": loan_number,
            "client_name": client_name,
            "days_overdue": days_late,
            "balance": balance,
            "sms_status": None,
            "email_status": None,
        }

        # Send SMS
        if body.method in ("sms", "both") and client_phone and _notification_service:
            try:
                if hasattr(_notification_service, 'twilio_client') and _notification_service.twilio_client:
                    _notification_service.twilio_client.messages.create(
                        body=sms_msg,
                        from_=_notification_service.twilio_phone,
                        to=client_phone
                    )
                    results["sms_sent"] += 1
                    notification_record["sms_status"] = "sent"
                else:
                    results["sms_failed"] += 1
                    notification_record["sms_status"] = "no_twilio"
            except Exception as e:
                logger.error(f"SMS error for {client_name}: {e}")
                results["sms_failed"] += 1
                notification_record["sms_status"] = f"error: {str(e)[:50]}"

        # Send Email
        if body.method in ("email", "both") and client_email and _notification_service:
            try:
                if hasattr(_notification_service, 'sendgrid_client') and _notification_service.sendgrid_client:
                    from sendgrid.helpers.mail import Mail, Email, To, Content
                    email_html = f"""
                    <div style="font-family: Arial, sans-serif; max-width: 500px; margin: 0 auto;">
                        <div style="background: #8B1513; padding: 20px; text-align: center; border-radius: 8px 8px 0 0;">
                            <h2 style="color: white; margin: 0;">Ross Lending Solutions</h2>
                            <p style="color: rgba(255,255,255,0.8); margin: 5px 0 0 0;">Recordatorio de Pago</p>
                        </div>
                        <div style="padding: 25px; border: 1px solid #eee; border-top: 0;">
                            <p>Estimado/a <strong>{client_name}</strong>,</p>
                            <p>Le recordamos que su pago mensual de <strong>${monthly_payment:,.2f}</strong> 
                               del préstamo <strong>#{loan_number}</strong> se encuentra vencido 
                               hace <strong>{days_late} día(s)</strong>.</p>
                            <div style="background: #fff3f3; border: 1px solid #ffcccc; padding: 15px; border-radius: 8px; margin: 15px 0;">
                                <p style="margin: 0;"><strong>Balance actual:</strong> ${balance:,.2f}</p>
                                <p style="margin: 5px 0 0 0;"><strong>Pago mínimo:</strong> ${monthly_payment:,.2f}</p>
                            </div>
                            <p>Por favor comuníquese con nosotros al <strong>(806) 503-7721</strong> para realizar su pago.</p>
                            <p style="font-size: 12px; color: #888;">Este es un recordatorio automático de Ross Lending Solutions LLC.</p>
                        </div>
                    </div>
                    """
                    mail = Mail(
                        from_email=Email(_notification_service.from_email),
                        to_emails=To(client_email),
                        subject=f"⚠️ Recordatorio de Pago — Préstamo #{loan_number}",
                        html_content=Content("text/html", email_html)
                    )
                    _notification_service.sendgrid_client.send(mail)
                    results["email_sent"] += 1
                    notification_record["email_status"] = "sent"
                else:
                    results["email_failed"] += 1
                    notification_record["email_status"] = "no_sendgrid"
            except Exception as e:
                logger.error(f"Email error for {client_name}: {e}")
                results["email_failed"] += 1
                notification_record["email_status"] = f"error: {str(e)[:50]}"

        results["details"].append(notification_record)

        # Log notification in DB
        await db["collection_notifications"].insert_one({
            "loan_id": str(loan["_id"]),
            "loan_number": loan_number,
            "client_name": client_name,
            "client_phone": client_phone,
            "client_email": client_email,
            "method": body.method,
            "message": sms_msg,
            "days_overdue": days_late,
            "balance": balance,
            "sms_status": notification_record.get("sms_status"),
            "email_status": notification_record.get("email_status"),
            "sent_at": now.isoformat(),
            "sent_by": user.get("email", "admin"),
        })

        # Update loan: mark that notification was sent
        await db[COLLECTION_NAME].update_one(
            {"_id": loan["_id"]},
            {"$set": {
                "last_notification_date": now.isoformat(),
                "last_notification_method": body.method,
                "status": "delinquent" if loan.get("status") != "delinquent" and days_late > 30 else loan.get("status"),
            }}
        )

    return {
        "success": True,
        "total_notified": len(overdue_loans),
        "sms_sent": results["sms_sent"],
        "sms_failed": results["sms_failed"],
        "email_sent": results["email_sent"],
        "email_failed": results["email_failed"],
        "details": results["details"],
    }


@regulated_lender_router.get("/admin/regulated-loans/collection-history")
async def get_collection_history(request: Request, limit: int = 50):
    """Get history of collection notifications sent."""
    await _auth_admin(request)
    db = _db

    notifications = []
    async for notif in db["collection_notifications"].find().sort("sent_at", -1).limit(limit):
        notif["_id"] = str(notif["_id"])
        notifications.append(notif)

    return {"notifications": notifications, "total": len(notifications)}

@regulated_lender_router.get("/admin/regulated-loans/{loan_id}")
async def get_loan(request: Request, loan_id: str):
    await _auth_admin(request)
    db = _db
    loan = await db[COLLECTION_NAME].find_one({"_id": ObjectId(loan_id)})
    if not loan:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    loan["_id"] = str(loan["_id"])
    return loan


@regulated_lender_router.put("/admin/regulated-loans/{loan_id}")
async def update_loan(request: Request, loan_id: str):
    user = await _auth_admin(request)
    db = _db
    body = await request.json()
    body["updated_at"] = datetime.now().isoformat()
    body.pop("_id", None)
    await db[COLLECTION_NAME].update_one({"_id": ObjectId(loan_id)}, {"$set": body})
    return {"success": True}


@regulated_lender_router.delete("/admin/regulated-loans/{loan_id}")
async def delete_loan(request: Request, loan_id: str):
    await _auth_admin(request)
    db = _db
    await db[COLLECTION_NAME].delete_one({"_id": ObjectId(loan_id)})
    await db[PAYMENTS_COLLECTION].delete_many({"loan_id": loan_id})
    return {"success": True}


@regulated_lender_router.post("/admin/regulated-loans/{loan_id}/evaluate-credit")
async def evaluate_credit(request: Request, loan_id: str):
    """
    Manual credit evaluation — Admin enters credit data, system runs it through
    the approval rules engine and updates the loan status accordingly.
    """
    user = await _auth_admin(request)
    db = _db
    body = await request.json()

    loan = await db[COLLECTION_NAME].find_one({"_id": ObjectId(loan_id)})
    if not loan:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    # Build applicant credit data from manual input
    credit_data = {
        "credit_score": int(body.get("credit_score", 0)),
        "dti": float(body.get("dti", 0)),
        "has_bankruptcy": bool(body.get("has_bankruptcy", False)),
        "collections_count": int(body.get("collections_count", 0)),
        "monthly_income": float(body.get("monthly_income", 0)),
        "employment_months": int(body.get("employment_months", 0)),
        "existing_debt": float(body.get("existing_debt", 0)),
        "source": body.get("source", "manual"),  # manual | microbilt | equifax | transunion
        "notes": body.get("notes", ""),
    }

    # Get active approval rules
    from approval_engine_router import evaluate_application, RULES_COLLECTION
    rules_cursor = db[RULES_COLLECTION].find({"active": True})
    rules = await rules_cursor.to_list(length=100)

    # If no rules in DB, use defaults
    if not rules:
        from approval_engine_router import DEFAULT_RULES
        rules = [r for r in DEFAULT_RULES if r.get("active", True)]

    # Run evaluation
    result = evaluate_application(rules, credit_data)

    # Map decision to loan status
    decision = result.get("decision", "manual_review")
    status_map = {
        "auto_approve": "pending_signature",  # Approved → ready to sign
        "manual_review": "under_review",       # Needs manual admin review
        "auto_decline": "declined",            # Rejected
    }
    new_status = status_map.get(decision, "under_review")

    # Save credit data and evaluation to the loan
    evaluation_record = {
        "credit_data": credit_data,
        "evaluation_result": {
            "decision": decision,
            "decision_label": result.get("decision_label", ""),
            "matched_rule": result.get("matched_rule"),
            "max_amount": result.get("max_amount", 0),
            "suggested_rate_min": result.get("suggested_rate_min", 0),
            "suggested_rate_max": result.get("suggested_rate_max", 0),
            "decline_reason": result.get("decline_reason", ""),
            "notes": result.get("notes", ""),
        },
        "evaluated_at": datetime.now().isoformat(),
        "evaluated_by": user.get("email", "admin"),
    }

    update_doc = {
        "credit_data": credit_data,
        "credit_evaluation": evaluation_record,
        "status": new_status,
        "updated_at": datetime.now().isoformat(),
    }

    await db[COLLECTION_NAME].update_one(
        {"_id": ObjectId(loan_id)},
        {"$set": update_doc}
    )

    # Also save to evaluation history
    eval_history = {
        "loan_id": loan_id,
        "loan_number": loan.get("loan_number", ""),
        "client_name": loan.get("client_name", ""),
        "amount": loan.get("amount", 0),
        **evaluation_record,
    }
    await db["approval_evaluations"].insert_one(eval_history)

    logger.info(f"Credit evaluation for {loan.get('loan_number')}: {decision} by {user.get('email')}")

    return {
        "success": True,
        "decision": decision,
        "decision_label": result.get("decision_label"),
        "matched_rule": result.get("matched_rule"),
        "new_status": new_status,
        "max_amount": result.get("max_amount", 0),
        "suggested_rate_min": result.get("suggested_rate_min", 0),
        "suggested_rate_max": result.get("suggested_rate_max", 0),
        "decline_reason": result.get("decline_reason", ""),
        "credit_data": credit_data,
    }


@regulated_lender_router.post("/admin/regulated-loans/{loan_id}/payments")
async def record_payment(request: Request, loan_id: str, body: PaymentCreate):
    user = await _auth_admin(request)
    db = _db

    loan = await db[COLLECTION_NAME].find_one({"_id": ObjectId(loan_id)})
    if not loan:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    # Record payment
    payment_count = await db[PAYMENTS_COLLECTION].count_documents({"loan_id": loan_id})
    payment_doc = {
        "loan_id": loan_id,
        "payment_number": payment_count + 1,
        "amount": body.amount,
        "payment_method": body.payment_method,
        "source": body.source,
        "notes": body.notes,
        "payment_date": datetime.now().isoformat(),
        "recorded_by": user.get("email", "admin"),
    }
    await db[PAYMENTS_COLLECTION].insert_one(payment_doc)

    # Update loan balance
    new_balance = max(0, loan.get("balance", 0) - body.amount)
    # Allocate: fees first, then interest, then principal
    remaining = body.amount
    fees_paid_increment = 0
    interest_paid_increment = 0
    principal_paid_increment = 0

    outstanding_fees = loan.get("admin_fee", 0) - loan.get("fees_paid", 0)
    if remaining > 0 and outstanding_fees > 0:
        fees_portion = min(remaining, outstanding_fees)
        fees_paid_increment = fees_portion
        remaining -= fees_portion

    outstanding_interest = loan.get("total_interest", 0) - loan.get("interest_paid", 0)
    if remaining > 0 and outstanding_interest > 0:
        int_portion = min(remaining, outstanding_interest)
        interest_paid_increment = int_portion
        remaining -= int_portion

    if remaining > 0:
        principal_paid_increment = remaining

    update_data = {
        "balance": round(new_balance, 2),
        "principal_paid": round(loan.get("principal_paid", 0) + principal_paid_increment, 2),
        "interest_paid": round(loan.get("interest_paid", 0) + interest_paid_increment, 2),
        "fees_paid": round(loan.get("fees_paid", 0) + fees_paid_increment, 2),
        "updated_at": datetime.now().isoformat(),
    }

    # If fully paid
    if new_balance <= 0:
        update_data["status"] = "paid_off"
        update_data["closed_date"] = datetime.now().isoformat()

    # If tax advance refund received
    if body.source == "refund_transfer":
        update_data["refund_status"] = "received"
        update_data["refund_received_date"] = datetime.now().isoformat()
        update_data["refund_amount_received"] = body.amount

    await db[COLLECTION_NAME].update_one({"_id": ObjectId(loan_id)}, {"$set": update_data})

    return {"success": True, "new_balance": new_balance, "paid_off": new_balance <= 0}


@regulated_lender_router.get("/admin/regulated-loans/{loan_id}/payments")
async def list_payments(request: Request, loan_id: str):
    await _auth_admin(request)
    db = _db
    payments = []
    async for p in db[PAYMENTS_COLLECTION].find({"loan_id": loan_id}).sort("payment_date", -1):
        p["_id"] = str(p["_id"])
        payments.append(p)
    return {"payments": payments}



# ═══════════════════════════════════════════════════════════════════════════════
# PAYMENT RECEIPT PDF GENERATION
# ═══════════════════════════════════════════════════════════════════════════════
@regulated_lender_router.get("/admin/regulated-loans/{loan_id}/payments/{payment_number}/receipt")
async def generate_payment_receipt(request: Request, loan_id: str, payment_number: int):
    """Generate a PDF receipt for a specific payment."""
    await _auth_admin(request)
    db = _db

    loan = await db[COLLECTION_NAME].find_one({"_id": ObjectId(loan_id)})
    if not loan:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    payment = await db[PAYMENTS_COLLECTION].find_one({"loan_id": loan_id, "payment_number": payment_number})
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.75*inch, bottomMargin=0.75*inch)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('ReceiptTitle', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#1a5632'), spaceAfter=6)
        subtitle_style = ParagraphStyle('ReceiptSubtitle', parent=styles['Normal'], fontSize=10, textColor=colors.grey, spaceAfter=20)
        heading_style = ParagraphStyle('ReceiptHeading', parent=styles['Heading3'], fontSize=12, textColor=colors.HexColor('#1a5632'), spaceBefore=16, spaceAfter=8)
        normal = styles['Normal']

        elements = []

        # Header
        elements.append(Paragraph("ROSS LENDING SOLUTIONS LLC", title_style))
        elements.append(Paragraph("Comprobante de Pago — Recibo Oficial", subtitle_style))

        # Receipt info
        pay_date = payment.get("payment_date", "")[:10] if payment.get("payment_date") else "N/A"
        receipt_data = [
            ["Número de Recibo:", f"RLS-PAY-{loan.get('loan_number', 'N/A')}-{payment_number:03d}"],
            ["Fecha de Pago:", pay_date],
            ["Registrado por:", payment.get("recorded_by", "admin")],
        ]
        t = Table(receipt_data, colWidths=[2*inch, 4*inch])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 12))

        # Client & Loan Info
        elements.append(Paragraph("Información del Préstamo", heading_style))
        loan_data = [
            ["Cliente:", loan.get("client_name", "N/A")],
            ["Teléfono:", loan.get("client_phone", "N/A")],
            ["Número de Préstamo:", loan.get("loan_number", "N/A")],
            ["Tipo:", loan.get("loan_type", "N/A").replace("_", " ").title()],
            ["Monto Original:", f"${loan.get('amount', 0):,.2f}"],
        ]
        t2 = Table(loan_data, colWidths=[2*inch, 4*inch])
        t2.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t2)
        elements.append(Spacer(1, 12))

        # Payment Details
        elements.append(Paragraph("Detalle del Pago", heading_style))
        method_map = {"cash": "Efectivo", "transfer": "Transferencia", "check": "Cheque", "card": "Tarjeta", "ach": "ACH"}
        pay_details = [
            ["Pago #:", str(payment_number)],
            ["Monto del Pago:", f"${payment.get('amount', 0):,.2f}"],
            ["Método de Pago:", method_map.get(payment.get("payment_method", ""), payment.get("payment_method", "N/A"))],
            ["Fuente:", payment.get("source", "manual")],
            ["Notas:", payment.get("notes", "—") or "—"],
        ]
        t3 = Table(pay_details, colWidths=[2*inch, 4*inch])
        t3.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#1a5632')),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0fdf4')),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#bbf7d0')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(t3)
        elements.append(Spacer(1, 16))

        # Balance Summary
        elements.append(Paragraph("Resumen de Cuenta", heading_style))
        balance_data = [
            ["Balance Actual:", f"${loan.get('balance', 0):,.2f}"],
            ["Principal Pagado:", f"${loan.get('principal_paid', 0):,.2f}"],
            ["Interés Pagado:", f"${loan.get('interest_paid', 0):,.2f}"],
            ["Fees Pagados:", f"${loan.get('fees_paid', 0):,.2f}"],
        ]
        t4 = Table(balance_data, colWidths=[2*inch, 4*inch])
        t4.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t4)
        elements.append(Spacer(1, 30))

        # Footer
        elements.append(Paragraph(
            "Este recibo es un comprobante oficial de pago emitido por Ross Lending Solutions LLC. "
            "Conserve este documento para sus registros. Para preguntas, contacte a su asesor financiero.",
            ParagraphStyle('Footer', parent=normal, fontSize=8, textColor=colors.grey, alignment=1)
        ))
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(
            f"Ross Lending Solutions LLC — OCCC Regulated Lender — Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            ParagraphStyle('FooterSmall', parent=normal, fontSize=7, textColor=colors.lightgrey, alignment=1)
        ))

        doc.build(elements)
        buf.seek(0)
        filename = f"Recibo_Pago_{loan.get('loan_number','RLS')}_{payment_number:03d}.pdf"
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando recibo: {str(e)}")


# ═══════════════════════════════════════════════════════════════════════════════
# TODAY'S PAYMENTS — DAILY RECONCILIATION
# ═══════════════════════════════════════════════════════════════════════════════
@regulated_lender_router.get("/admin/payments/today")
async def get_todays_payments(request: Request, date: str = "", license_type: str = "regulated"):
    """Get all payments recorded today (or a specific date) for daily reconciliation."""
    await _auth_admin(request)
    db = _db

    target_date = date or datetime.now().strftime("%Y-%m-%d")
    
    # Select collection based on license type
    if license_type == "cab":
        payments_col = "cab_payments"
        loans_col = "cab_loans"
    else:
        payments_col = PAYMENTS_COLLECTION
        loans_col = COLLECTION_NAME

    # Find all payments for that date
    payments = []
    async for p in db[payments_col].find().sort("payment_date" if license_type != "cab" else "created_at", -1):
        date_field = p.get("payment_date") or p.get("created_at", "")
        raw_date = date_field
        if hasattr(raw_date, 'strftime'):
            pay_date = raw_date.strftime("%Y-%m-%d")
        else:
            pay_date = str(raw_date)[:10]
        if pay_date == target_date:
            # Get loan info
            loan = None
            try:
                loan = await db[loans_col].find_one({"_id": ObjectId(p.get("loan_id", ""))})
            except:
                pass
            p["_id"] = str(p["_id"])
            p["client_name"] = p.get("client_name") or (loan.get("client_name", "N/A") if loan else "N/A")
            p["loan_number"] = p.get("loan_number") or (loan.get("loan_number", "N/A") if loan else "N/A")
            p["loan_balance"] = loan.get("balance", loan.get("remaining_balance", 0)) if loan else 0
            p["payment_date"] = p.get("payment_date") or p.get("created_at", "")
            p["recorded_by"] = p.get("recorded_by") or p.get("recorded_by_name", "")
            payments.append(p)

    total_collected = sum(p.get("amount", 0) for p in payments)
    by_method = {}
    for p in payments:
        m = p.get("payment_method", "other")
        if m not in by_method:
            by_method[m] = {"count": 0, "total": 0}
        by_method[m]["count"] += 1
        by_method[m]["total"] += p.get("amount", 0)

    return {
        "date": target_date,
        "payments": payments,
        "total_count": len(payments),
        "total_collected": round(total_collected, 2),
        "by_method": by_method,
    }


@regulated_lender_router.get("/admin/regulated-loans/{loan_id}/schedule")
async def get_schedule(request: Request, loan_id: str):
    await _auth_admin(request)
    db = _db
    loan = await db[COLLECTION_NAME].find_one({"_id": ObjectId(loan_id)})
    if not loan:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    schedule = generate_regulated_schedule(
        loan["amount"],
        loan.get("total_interest", 0),
        loan.get("admin_fee", 0),
        loan.get("term_months", 1),
        loan.get("first_payment_date"),
    )
    return {"schedule": schedule}


@regulated_lender_router.put("/admin/regulated-loans/{loan_id}/status")
async def update_status(request: Request, loan_id: str, body: StatusUpdate):
    user = await _auth_admin(request)
    db = _db

    loan = await db[COLLECTION_NAME].find_one({"_id": ObjectId(loan_id)})
    if not loan:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    history_entry = {
        "status": body.status,
        "date": datetime.now().isoformat(),
        "by": user.get("email", "admin"),
        "comment": body.comment,
    }

    await db[COLLECTION_NAME].update_one(
        {"_id": ObjectId(loan_id)},
        {
            "$set": {"status": body.status, "updated_at": datetime.now().isoformat()},
            "$push": {"status_history": history_entry},
        }
    )

    # ═══ TRIGGER EMAIL ON STATUS CHANGE (Background) ═══
    try:
        import asyncio
        from email_sender import send_loan_approved, send_loan_denied
        client_email = loan.get('client_email', '')
        client_name = loan.get('client_name', '')

        if body.status == 'approved' and client_email:
            asyncio.create_task(send_loan_approved(
                client_email=client_email,
                client_name=client_name,
                loan_number=loan.get('loan_number', ''),
                amount=loan.get('amount', 0),
                monthly_payment=loan.get('monthly_payment', 0),
                term_months=loan.get('term_months', 0)
            ))
        elif body.status == 'rejected' and client_email:
            asyncio.create_task(send_loan_denied(
                client_email=client_email,
                client_name=client_name
            ))
    except Exception as email_err:
        logger.error(f"Status change email error: {email_err}")

    return {"success": True}




# ═══════════════════════════════════════════════════════════════════════════════
# CONTRACT / DOCUMENT PDF GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

@regulated_lender_router.get("/admin/regulated-loans/{loan_id}/contract")
async def generate_contract_pdf(request: Request, loan_id: str, type: str = Query("loan_agreement"), lang: str = Query("es")):
    """
    Generate PDF contract documents for a regulated loan.
    Types: loan_agreement, truth_in_lending, payment_schedule, promissory_note
    Lang: 'es' (Spanish) or 'en' (English)
    """
    await _auth_admin(request)
    db = _db

    loan = await db[COLLECTION_NAME].find_one({"_id": ObjectId(loan_id)})
    if not loan:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    loan["_id"] = str(loan["_id"])

    if lang not in ("en", "es"):
        lang = "es"

    try:
        from loan_pdf_service import generate_loan_contract_pdf

        # Use shared service for schedule + PDF data mapping
        schedule = generate_schedule_for_loan(loan)
        pdf_loan = build_pdf_loan_data(loan)

        pdf_base64 = generate_loan_contract_pdf(pdf_loan, schedule, lang=lang)
        pdf_bytes = base64.b64decode(pdf_base64)

        # Determine filename based on type and language
        type_labels_es = {
            "loan_agreement": "Contrato",
            "truth_in_lending": "TIL_Divulgacion",
            "payment_schedule": "Amortizacion",
            "promissory_note": "Pagare",
        }
        type_labels_en = {
            "loan_agreement": "Loan_Agreement",
            "truth_in_lending": "Truth_In_Lending",
            "payment_schedule": "Amortization",
            "promissory_note": "Promissory_Note",
        }
        labels = type_labels_en if lang == "en" else type_labels_es
        label = labels.get(type, "Document" if lang == "en" else "Documento")
        lang_tag = lang.upper()
        filename = f"{label}_{loan.get('loan_number', 'RLS')}_{lang_tag}.pdf"

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(pdf_bytes)),
            }
        )
    except Exception as e:
        logger.error(f"Error generating contract PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")
